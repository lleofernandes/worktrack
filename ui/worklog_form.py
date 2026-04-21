"""
worklog_form.py — Formulário de controle de horas (por contrato).
"""
from __future__ import annotations

from datetime import date, time
from decimal import Decimal
import io

import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from fpdf import FPDF

from database.connection import SessionLocal
from database.models import ContractType
from database.repository import (
    ContractRepository,
    ContractRateRepository,
    ProjectRepository,
    WorkLogRepository,
)
from utils.calculations import calc_worked_hours
from utils.toast_helper import set_toast, show_pending_toast


CONTRACT_LABELS = {
    "WORK_HOUR": "Por Hora",
    "PROJECT": "Projeto Fechado",
    "PROJECT_HOURS": "Projeto c/ Horas",
}


def render_worklog_form() -> None:
    st.header("⏱️ Controle de Horas")
    show_pending_toast()

    session = SessionLocal()
    try:
        tab_hist, tab_form = st.tabs(["📋 Histórico", "➕ Adicionar"])
        with tab_hist:
            _render_history(session)

        with tab_form:
            _render_form(session)
    finally:
        session.close()


def _render_form(session) -> None:
    contracts = ContractRepository.get_all(session, active_only=True)
    if not contracts:
        st.warning("Nenhum contrato ativo cadastrado. Cadastre uma empresa e um contrato primeiro.")
        return

    contract_options = {
        f"[{ct.id}] {ct.company.name if ct.company else '?'} — {ct.contract_number or 'S/N'} ({CONTRACT_LABELS.get(ct.contract_type, ct.contract_type)})": ct.id
        for ct in contracts
    }

    col1, col2 = st.columns(2)
    with col1:
        selected_label = st.selectbox("Contrato *", list(contract_options.keys()), key="wl_contract")
        contract_id = contract_options[selected_label]
    with col2:
        log_date = st.date_input("Data *", value=date.today(), format="DD/MM/YYYY", key="wl_date")

    contract = ContractRepository.get_by_id(session, contract_id)
    rate_obj = ContractRateRepository.get_active_rate(session, contract_id, log_date)
    contract_type = contract.contract_type if contract else ContractType.WORK_HOUR

    with st.container(border=True):
        ci1, ci2, ci3, ci4 = st.columns(4)

        def field(label, value):
            st.markdown(
                f"""
                <div style="line-height:2.1">
                    <div style="font-size:14px; text-align: center; color:gray; padding: 0;">{label}</div>
                    <div style="font-size:18px; text-align: center; padding: 0 0 5px 0; ">{value}</div>
                </div>
                """,
                unsafe_allow_html=True,
            )

        with ci1:
            field("Empresa", contract.company.fantasy_name or contract.company.name if contract and contract.company else "—")

        with ci2:
            field("Tipo de Contrato", CONTRACT_LABELS.get(contract_type, contract_type))

        with ci3:
            field("Nº Contrato", contract.contract_number or "—" if contract else "—")

        with ci4:
            field("Taxa/h", f"R$ {float(rate_obj.hour_rate):,.2f}" if rate_obj else "⚠️ Sem taxa")

    projects = ProjectRepository.get_all_by_contract(session, contract_id)
    project_options = {"— Nenhum —": None}
    project_options.update({p.name: p.id for p in projects})
    selected_project = st.selectbox("Projeto (opcional)", list(project_options.keys()), key="wl_project")
    project_id = project_options[selected_project]

    st.divider()

    if contract_type == ContractType.WORK_HOUR:
        start_time, end_time, break_minutes, extra_partner_minutes, total_hours, progress_pct = _fields_work_hour()
    elif contract_type == ContractType.PROJECT_HOURS:
        start_time, end_time, break_minutes, extra_partner_minutes, total_hours, progress_pct = _fields_project_hours()
    else:
        start_time, end_time, break_minutes, extra_partner_minutes, total_hours, progress_pct = _fields_project()

    description = st.text_area(
        "Descrição / Atividades",
        height=100,
        key="wl_desc",
        placeholder="Descreva as atividades realizadas...",
    )

    preview_hours = _preview_hours(
        contract_type,
        start_time,
        end_time,
        break_minutes,
        extra_partner_minutes,
        total_hours,
    )

    if preview_hours and preview_hours > 0:
        rev = float(preview_hours) * float(rate_obj.hour_rate) if rate_obj and contract_type == ContractType.WORK_HOUR else 0
        pc1, pc2 = st.columns(2)
        pc1.info(f"⏱️ **{float(preview_hours):.2f}h calculadas**")
        if rate_obj and contract_type == ContractType.WORK_HOUR:
            pc2.info(f"💰 **R$ {rev:,.2f} estimado**")

    progress_pct = None
    if contract_type == ContractType.PROJECT_HOURS:
        logs_existentes = WorkLogRepository.list_by_contract_month(
            session,
            contract_id,
            log_date.year,
            log_date.month,
        )
        horas_acumuladas = sum(float(wl.total_hours or 0) for wl in logs_existentes)
        horas_contratadas = float(contract.contracted_hours or 0) if contract else 0

        if horas_contratadas > 0:
            horas_com_preview = horas_acumuladas + float(preview_hours or 0)
            progress_pct = int(round((horas_com_preview / horas_contratadas) * 100, 0))
            st.info(f"📈 **Progress:** {progress_pct}% do pacote de horas")
        else:
            progress_pct = None

    if st.button("💾 Salvar Apontamento", type="primary", width="stretch", key="wl_save"):
        error = _validate(contract_type, start_time, end_time, break_minutes, total_hours, progress_pct, preview_hours)
        if error:
            st.error(f"❌ {error}")
        else:
            try:
                descricao_normalizada = (description or "").strip()
                logs_mesmo_periodo = WorkLogRepository.list_by_contract_month(
                    session,
                    contract_id,
                    log_date.year,
                    log_date.month,
                )

                duplicado = next(
                    (
                        wl for wl in logs_mesmo_periodo
                        if wl.date == log_date
                        and (wl.project_id or None) == (project_id or None)
                        and (wl.description or "").strip() == descricao_normalizada
                    ),
                    None,
                )

                if duplicado:
                    st.error("❌ Já existe um apontamento com a mesma Data, Contrato, Projeto e Descrição.")
                    return

                total_h = None
                if start_time and end_time:
                    total_h = Decimal(str(calc_worked_hours(
                        start_time,
                        end_time,
                        break_minutes or 0,
                        extra_partner_minutes or 0,
                    )))
                elif total_hours is not None:
                    total_h = Decimal(str(total_hours))

                wl = WorkLogRepository.create(
                    session,
                    contract_id=contract_id,
                    project_id=project_id or None,
                    date=log_date,
                    start_time=start_time,
                    end_time=end_time,
                    break_minutes=break_minutes or 0,
                    extra_partner_minutes=int(extra_partner_minutes or 0),
                    total_hours=total_h,
                    progress_pct=progress_pct if contract_type == ContractType.PROJECT_HOURS else None,
                    description=descricao_normalizada or None,
                )

                session.commit()
                label = f"{float(preview_hours):.2f}h" if preview_hours else "registro"
                set_toast(f"✅ Apontamento #{wl.id} salvo — {label} em {log_date.strftime('%d/%m/%Y')}!")
                st.rerun()
            except Exception as e:
                session.rollback()
                st.error(f"❌ Erro: {e}")


def _fields_work_hour():
    st.caption("⏰ **Por Hora** — informe o horário trabalhado")
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        start_time = st.time_input("Início *", value=time(9, 0), step=1800, key="wl_start")
    with c2:
        end_time = st.time_input("Término *", value=time(18, 0), step=1800, key="wl_end")
    with c3:
        break_min = st.number_input("Intervalo (min)", 0, 480, 60, 15, key="wl_break")
    with c4:
        extra_min = st.number_input(
            "Extra Parceiro (min)",
            min_value=0,
            max_value=480,
            value=0,
            step=5,
            key="wl_extra",
            help="Minutos extras do parceiro. Ex: 45, 90, 120...",
        )
    return start_time, end_time, break_min, extra_min, None, None


def _fields_project_hours():
    st.caption("📦 **Projeto c/ Horas** — total de horas do período")
    c1, c2 = st.columns([1, 2])
    with c1:
        total = st.number_input("Total de Horas *", 0.25, 24.0, 8.0, 0.25, format="%.2f", key="wl_total")
    with c2:
        h, m = int(total), int((total % 1) * 60)
        st.info(f"⏱️ **{h}h {m:02d}min** de trabalho no projeto.")
    return None, None, 0, 0.0, total, None


def _fields_project():
    st.caption("🎯 **Projeto Fechado** — registre as horas do período")
    c1, c2 = st.columns([1, 2])
    with c1:
        total = st.number_input("Total de Horas *", 0.25, 24.0, 8.0, 0.25, format="%.2f", key="wl_total_project")
    with c2:
        h, m = int(total), int((total % 1) * 60)
        st.info(f"⏱️ **{h}h {m:02d}min** de trabalho no projeto.")
    st.caption("ℹ️ Faturamento controlado pela tela de Notas Fiscais.")
    return None, None, 0, 0.0, total, None


def _preview_hours(ct, start, end, brk, extra, total):
    if ct == ContractType.WORK_HOUR and start and end:
        try:
            return calc_worked_hours(start, end, brk or 0, extra or 0)
        except Exception:
            return None

    if ct in (ContractType.PROJECT_HOURS, ContractType.PROJECT) and total:
        return Decimal(str(total))
    return None


def _validate(ct, start, end, brk, total, pct, preview):
    if ct == ContractType.WORK_HOUR:
        if not start or not end:
            return "Horário de início e término obrigatórios."
        if end <= start:
            return "Término deve ser maior que início."
        if preview and preview > 24:
            return "Horas não podem exceder 24h/dia."
        if preview and preview <= 0:
            return "Total de horas deve ser maior que zero."
    elif ct in (ContractType.PROJECT_HOURS, ContractType.PROJECT):
        if not total or total <= 0:
            return "Informe o total de horas (> 0)."
        if total > 24:
            return "Total não pode exceder 24h/dia."
    return None


def _render_history(session) -> None:
    st.subheader("📋 Histórico de Apontamentos")

    contracts = ContractRepository.get_all(session, active_only=None)
    if not contracts:
        st.info("Nenhum contrato cadastrado.")
        return

    with st.container(border=True):
        hf1, hf2, hf3, hf4 = st.columns(4)

        with hf1:
            sel_status = st.selectbox("Status Contrato", ["Todos", "Ativo", "Inativo"], key="hist_status")
            active_filter = {"Todos": None, "Ativo": True, "Inativo": False}[sel_status]

        with hf2:
            contracts_filtered = ContractRepository.get_all(session, active_only=active_filter)
            ct_opts = {"Todos": None}
            ct_opts.update({
                f"[{ct.id}] {ct.company.fantasy_name or ct.company.name} — {ct.contract_number or 'S/N'}": ct.id
                for ct in contracts_filtered
            })
            sel_ct = st.selectbox("Contrato", list(ct_opts.keys()), key="hist_contract")
            filter_ct_id = ct_opts[sel_ct]

        with hf3:
            year_opts = ["Todos"] + list(range(2021, date.today().year + 1))
            sel_year = st.selectbox(
                "Ano",
                year_opts,
                index=year_opts.index(date.today().year),
                key="hist_year",
            )
            filter_year = None if sel_year == "Todos" else int(sel_year)

        with hf4:
            month_opts = {"Todos": None}
            month_opts.update({
                date(2021, m, 1).strftime("%B").capitalize(): m
                for m in range(1, 13)
            })
            month_labels = list(month_opts.keys())
            cur_month_label = date(2021, date.today().month, 1).strftime("%B").capitalize()
            sel_month = st.selectbox(
                "Mes",
                month_labels,
                index=month_labels.index(cur_month_label),
                key="hist_month",
            )
            filter_month = month_opts[sel_month]

    logs = WorkLogRepository.list_filtered(
        session,
        contract_id=filter_ct_id,
        year=filter_year,
        month=filter_month,
    )

    if active_filter is not None:
        logs = [wl for wl in logs if wl.contract and wl.contract.is_active == active_filter]

    if not logs:
        st.info("Nenhum apontamento encontrado.")
        return

    rows = []
    for wl in logs:
        ct = wl.contract
        company = ct.company.fantasy_name or ct.company.name if ct and ct.company else "-"

        if wl.total_hours is not None and float(wl.total_hours) > 0:
            horas = float(wl.total_hours)
        elif wl.start_time and wl.end_time:
            horas = float(calc_worked_hours(
                wl.start_time,
                wl.end_time,
                wl.break_minutes or 0,
                wl.extra_partner_minutes or 0,
            ))
        else:
            horas = 0.0

        rows.append({
            "Data": wl.date.strftime("%d/%m/%Y"),
            "Cliente": company,
            "Contrato": ct.contract_number or "S/N" if ct else "-",
            "Projeto": wl.project.name if wl.project else "-",
            "Inicio": wl.start_time.strftime("%H:%M") if wl.start_time else "-",
            "Fim": wl.end_time.strftime("%H:%M") if wl.end_time else "-",
            "Intervalo(min)": wl.break_minutes or 0,
            "Horas": horas,
            "Progresso": float(wl.progress_pct) if wl.progress_pct is not None else None,
            "Descricao": wl.description or "",
        })

    df = pd.DataFrame(rows)
    df["Data"] = pd.to_datetime(df["Data"], dayfirst=True)
    df = df.sort_values(by="Data", ascending=True).reset_index(drop=True)
    df["Data"] = df["Data"].dt.strftime("%d/%m/%Y")

    total_horas = df["Horas"].sum()
    total_registros = len(df)

    mc1, mc2, mc3, mc4, mc5 = st.columns([2, 2, 2, 1.5, 1.5])

    mc1.metric("Total de Registros", total_registros)
    mc2.metric("Total de Horas", f"{total_horas:.2f}h")
    mc3.metric("Contratos", len(set(df["Contrato"].tolist())))

    with mc4:
        st.write("")
        excel_bytes = _export_excel(df, filter_year, filter_month)
        st.download_button(
            label="📥 Excel",
            data=excel_bytes,
            file_name=f"lf_relatorio_{filter_year or 'todos'}_{filter_month or 'todos'}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
        )

    with mc5:
        st.write("")
        pdf_bytes = _export_pdf(df, filter_year, filter_month, total_horas)
        st.download_button(
            label="📄 PDF",
            data=pdf_bytes,
            file_name=f"lf_relatorio_{filter_year or 'todos'}_{filter_month or 'todos'}.pdf",
            mime="application/pdf",
            use_container_width=True,
        )

    df["Progresso (%)"] = df["Progresso"].apply(lambda x: f"{x:.1f}%" if x is not None and float(x) > 0 else "—")

    st.dataframe(
        df.drop(columns=["Progresso"]),
        width="stretch",
        hide_index=True,
    )


# ── Gerador Excel ─────────────────────────────────────────────────────
def _export_excel(df, year, month) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Apontamentos"

    header_fill = PatternFill("solid", fgColor="01696F")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    periodo = f"{month or 'Todos'}/{year or 'Todos'}"
    ws.merge_cells("A1:J1")
    ws["A1"] = f"Relatório de Apontamentos — Período: {periodo}"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:J2")
    ws["A2"] = f"Gerado em: {date.today().strftime('%d/%m/%Y')}"
    ws["A2"].font = Font(italic=True, size=10, color="888888")
    ws["A2"].alignment = Alignment(horizontal="center")

    cols = list(df.columns)
    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    alt_fill = PatternFill("solid", fgColor="F0F4F4")
    for row_idx, row in enumerate(df.itertuples(index=False), start=5):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            if fill:
                cell.fill = fill

    total_row = len(df) + 6
    ws.cell(row=total_row, column=1, value="TOTAL HORAS").font = Font(bold=True)
    ws.cell(row=total_row, column=8, value=df["Horas"].sum()).font = Font(bold=True)

    col_widths = [12, 25, 14, 20, 8, 8, 15, 8, 13, 40]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()



def _clean(val: str) -> str:
    return (str(val)
            .replace("—", "-")
            .replace("–", "-")
            .replace("\u2026", "...")
            .replace("\u201c", '"').replace("\u201d", '"')
            .replace("\u2018", "'").replace("\u2019", "'")
            .encode("latin-1", errors="replace")
            .decode("latin-1"))


class PDFWithFooter(FPDF):
    def __init__(self, periodo: str, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.periodo = periodo

    def footer(self):
        self.set_y(-10)
        self.set_font("Helvetica", "", 8)
        self.set_text_color(100, 100, 100)
        texto = f"Documento gerado em: {date.today().strftime('%d/%m/%Y')}     |     Período: {self.periodo}"
        self.cell(0, 6, texto, align="R")


# ── Gerador PDF ───────────────────────────────────────────────────────
def _export_pdf(df, year, month, total_horas: float) -> bytes:
    periodo = f"{month or 'Todos'}/{year or 'Todos'}"

    pdf = PDFWithFooter(periodo=periodo, orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 14)
    pdf.set_fill_color(1, 105, 111)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 10, "LF Analytics - Relatório de Apontamentos", ln=True, align="C", fill=True)

    pdf.set_font("Helvetica", "B", 9)
    pdf.set_text_color(1, 105, 111)
    pdf.cell(0, 6, f"Total de Horas: {total_horas:.2f}h  |  Total de Registros: {len(df)}", ln=True, align="C")
    pdf.ln(3)

    cols = ["Data", "Cliente", "Contrato", "Projeto", "Horas", "Descrição"]
    widths = [25, 65, 25, 50, 25, 87]

    pdf.set_font("Helvetica", "B", 8)
    pdf.set_fill_color(1, 105, 111)
    pdf.set_text_color(255, 255, 255)
    for col, w in zip(cols, widths):
        pdf.cell(w, 7, col, border=1, align="C", fill=True)
    pdf.ln()

    pdf.set_font("Helvetica", "", 7.5)
    pdf.set_text_color(30, 30, 30)

    for i, row in df.iterrows():
        fill = i % 2 == 0
        pdf.set_fill_color(240, 248, 248) if fill else pdf.set_fill_color(255, 255, 255)

        desc = str(row.get("Descricao", "") or "")
        desc_fmt = (desc[:45] + "...") if len(desc) > 45 else desc
        horas_val = row.get("Horas", 0) or 0

        values = [
            _clean(str(row.get("Data", ""))),
            _clean(str(row.get("Cliente", ""))),
            _clean(str(row.get("Contrato", ""))),
            _clean(str(row.get("Projeto", ""))),
            _clean(f"{float(horas_val):.2f}"),
            _clean(desc_fmt),
        ]
        for val, w in zip(values, widths):
            pdf.cell(w, 6, str(val), border=1, align="C", fill=fill)
        pdf.ln()

    return bytes(pdf.output())
